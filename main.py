import PyPDF2
import re
import os
import tkinter as tk
from tkinter import filedialog, ttk
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
import string
from time import sleep


class Application:
    def __init__(self, master):
        self.database_numeros = {}
        self.locais = ['Sureg Arapoti PR/SP', 'AG Arapoti PR', 'AG Arapoti Centro PR', 'AG Jaguariaíva PR',
                  'AG Senges PR', 'AG Itararé SP', 'AG Itapeva SP', 'AG Itapeva Jd Maringá SP', 'AG Capão Bonito SP',
                  'AG Burí SP', 'AG Barão de Antonina SP', 'AG Taquarituba SP', 'AG Fartura SP', 'AG Riversul SP',
                  'AG Taguaí SP', 'AG Coronel Macedo SP', 'AG Itaí', 'AG Itaberá', 'AG Itaporanga', 'AG Ribeirão', 
                  'AG Taquarivai', 'Ainda não definido']

        self.sheet_database = './Lista.xlsx'

        self.master = master
        self.master.title("Rateio de linhas Vivo - Gustavo")
        self.master.geometry("400x150")
        self.master.resizable(False, False)
        self.master.configure(bg='#F0F0F0') # define a cor de fundo da janela

        # define o estilo dos widgets usando ttk
        style = ttk.Style(self.master)
        style.configure('TLabel', background='#F0F0F0', font=('Calibri', 12))
        style.configure('TButton', font=('Arial', 10))

        self.label = ttk.Label(self.master, text="Selecione um arquivo PDF para extrair números:")
        self.label.pack(anchor="center", padx=10, pady=12)

        # cria um novo Frame e empacota os botões nele
        self.button_frame = ttk.Frame(self.master)
        self.button_frame.pack(anchor='center', padx=10, pady=10)

        self.btn_font_file = ttk.Button(self.button_frame, text="Seleciona a Fonte", command=self.select_fonte, padding=(5))
        self.btn_font_file.pack(side=tk.LEFT, padx=(0,5))

        self.btn_select_file = ttk.Button(self.button_frame, text="Selecionar a Fatura", command=self.select_file, padding=(5))
        self.btn_select_file.config(state="disabled")
        self.btn_select_file.pack(side=tk.LEFT, padx=(5,0))

        self.status_label = ttk.Label(self.master, text="")
        self.status_label.pack()

    def select_fonte(self):
        file_path = filedialog.askopenfilename()
        if file_path and '.xlsx' in str(file_path).lower():

            self.sheet_database = file_path

            style = ttk.Style()
            style.configure('Green.TButton', foreground='green')

            self.btn_font_file.configure(style='Green.TButton')

            self.btn_select_file.config(state="normal")
            self.master.update()
        else:
            style = ttk.Style()
            style.configure('Red.TButton', foreground='red')

            self.btn_select_file.configure(style='Red.TButton')

    def select_file(self):
        file_path = filedialog.askopenfilename()
        if file_path and '.pdf' in str(file_path).lower():

            style = ttk.Style()
            style.configure('Green.TButton', foreground='green')
            self.btn_select_file.configure(style='Green.TButton')
            self.status_label.config(text="Extraindo números...")
            self.master.update()

            try:
                self.load_numers_database()
                self.extract_numbers(file_path)
                self.open_result()

            except PermissionError:
                self.status_label.config(text="ERRO: Feche a planilha numeros.xlsx")
                self.master.update()
            except ValueError:
                self.status_label.config(text="ERRO: Fonte incorreta ou inexistente.")
                self.master.update()

        else:
            style = ttk.Style()
            style.configure('Red.TButton', foreground='red')

            self.btn_select_file.configure(style='Red.TButton')
        
    def extract_numbers(self, file_path):
        # Abre o arquivo PDF em modo leitura binária
        with open(file_path, 'rb') as pdf_file:
            # Cria um objeto PDFReader
            pdf_reader = PyPDF2.PdfReader(pdf_file)

            # Inicializa um dicionário para armazenar os dados extraídos
            data = {'Número': [], 'MB': [], 'KB': [], 'Local': [], 'Agência': [], 'Valor': []}
            internet_numeros = {}

            # Loop pelas páginas do PDF
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]

                # Extrai o texto da página
                text = page.extract_text()

                match_internet = re.findall(r'(\d{2}-\d{5}-\d{4})\s+(\d+)MB\s+(\d+)KB', text)
                for encontrado in match_internet:
                    telefone = encontrado[0].strip().replace('-', '')

                    internet_numeros[str(telefone)] = [int(encontrado[1]), int(encontrado[2])]

            # Loop pelas páginas do PDF
            for page_num in range(len(pdf_reader.pages)):
                page = pdf_reader.pages[page_num]

                # Extrai o texto da página
                text = page.extract_text()

                # Procura por informações usando a expressão regular
                pattern = r'\d{2}[-.\s]?\d{5}[-.\s]?\d{4}\s+SMART\s+EMPRESAS\s+\d+GB\s+[A-Z]\s+\d+,\d+'
                matches = re.findall(pattern, text)

                # Adiciona as informações encontradas ao dicionário
                for match in matches:
                    informacao = match.replace(" SMART EMPRESAS 5GB D ", ";")

                    numero = informacao[:informacao.find(";")]
                    valor = str(informacao[informacao.find(";")+1:]).strip()
                    agencia = "Ainda não definido"
                    local = '0753-01'

                    numero = numero.replace('-', '')

                    # Tentar encontrar o Local
                    try:
                        if str(numero).strip() in self.database_numeros and len(self.database_numeros[numero][0]) >= 1:
                            agencia = self.database_numeros[numero][0]
                            local = self.database_numeros[numero][1]
                        else:
                            local = "0753-98"
                            agencia = 'Ainda não definido'

                    except Exception:
                        pass

                    data['Número'].append(numero)
                    data['Agência'].append(agencia)

                    if str(numero) in internet_numeros:
                        data['MB'].append(internet_numeros[numero][0])
                        data['KB'].append(internet_numeros[numero][1])
                    else:
                        data['MB'].append(0)
                        data['KB'].append(0)

                    data['Local'].append(local)
                    data['Valor'].append(valor.replace(".", '').replace(',', '.'))

            # Cria um objeto DataFrame com os dados extraídos
            df = pd.DataFrame(data)
            df = df.astype({'Número': str, 'MB': int, 'KB': int, 'Agência': str, 'Local': str, 'Valor': float})

            with pd.ExcelWriter("numero.xlsx") as writer:
                df.to_excel(writer, sheet_name='Numeros', startrow=len(self.locais)+4,index=False)

            self.format_excel_sheet()

    def format_excel_sheet(self, path_name='numero.xlsx'):
        valores_celulas = {'B2': 'Qtd.Linhas', 'C2':'Cons.Internet (GB)', 'D2':'Valor'}
        
        worksheet = openpyxl.load_workbook(path_name)
        planilha = worksheet.active

        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000'),
        )

        # Formatar Tabela de cima (Tabela de Rateio)
        for celula, valor in valores_celulas.items():
            planilha[celula] = valor
            planilha[celula].border = border
            planilha[celula].fill = PatternFill(start_color="00b050", end_color="00b050", fill_type='solid')
            planilha[celula].font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            planilha[celula].alignment = Alignment(horizontal='center', vertical='center')

        for linha, valor in enumerate(self.locais):
            planilha[f'A{linha+3}'] = valor
            planilha[f'A{linha+3}'].fill = PatternFill(start_color="00b050", end_color="00b050", fill_type='solid')
            planilha[f'A{linha+3}'].font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            planilha[f'A{linha+3}'].alignment = Alignment(horizontal='center', vertical='center')

        
        cells_range = planilha[f"B3:D{len(self.locais)+2}"]
        for linha in cells_range:
            for celula in linha:
                if celula.value is None or len(celula.value) <= 0:
                    celula.value = int(0)

                celula.border = border
                celula.alignment = Alignment(horizontal='center', vertical='center')

        # Formatar Tabela de Números (Tabela de baixo)
        limite_horizontal = 0
        limite_horizontal_letra = 'A'

        # Localizando cabeçalho
        tamanho_x = len(self.locais)+5
        for letra in string.ascii_uppercase:
            if planilha[f"{letra}{tamanho_x}"].value is None or len(str(planilha[f"{letra}{tamanho_x}"].value)) <=0:
                break
            limite_horizontal +=1
            limite_horizontal_letra = letra

        # Localizando corpo da tabela
        limite_vertical = len(self.locais)+4
        while True:
            limite_vertical +=1
            if planilha[f"A{limite_vertical}"].value is None or len(planilha[f"A{limite_vertical}"].value) <=0:
                break
            else:
                for letra in string.ascii_uppercase[:limite_horizontal]:
                    planilha[f"{letra}{limite_vertical}"].alignment = Alignment(horizontal='center', vertical='center')
                    planilha[f"{letra}{limite_vertical}"].border = border

        # Transformando em tabela
        tabela_rateio = Table(displayName='TabelaNumeros', ref=f'A{len(self.locais)+5}:{limite_horizontal_letra}{limite_vertical-1}')
        tabela_rateio.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                       showLastColumn=False, showRowStripes=True, showColumnStripes=True)

        planilha.add_table(tabela_rateio)

        # Formatando cabeçalho para verde
        for letra in string.ascii_uppercase[:limite_horizontal]:
            planilha[f"{letra}{len(self.locais)+5}"].fill = PatternFill(start_color="00b050", end_color="00b050", fill_type='solid')
            planilha[f"{letra}{len(self.locais)+5}"].font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
            planilha[f"{letra}{len(self.locais)+5}"].alignment = Alignment(horizontal='center', vertical='center')

        # Formatando corpo para preto e branco
        for letra in string.ascii_uppercase[:limite_horizontal]:
            for celula in range(len(self.locais)+6, limite_vertical):
                planilha[f"{letra}{celula}"].alignment = Alignment(horizontal='center', vertical='center')
                planilha[f"{letra}{celula}"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type='solid')
                
        # Ajustando automaticamente a largura colunas
        for letra in string.ascii_uppercase[:limite_horizontal]:
            max_length = 0

            for linha in range(1, limite_vertical):
                if planilha[f'{letra}{linha}'].value is not None and len(str(planilha[f'{letra}{linha}'].value)) > max_length:
                    max_length = len(str(planilha[f'{letra}{linha}'].value))
            
            planilha.column_dimensions[letra].width = max_length+2
        
        # Salvar planilha 
        worksheet.save(path_name)
        self.status_label.config(text="Extração concluída com sucesso!")

        self.comiple_info(limite_horizontal_letra, limite_vertical, path_name)

    def load_numers_database(self, file_path='./Lista.xlsx'):
        df = pd.read_excel(self.sheet_database, sheet_name='Relação Números')
        for index, row in df.iterrows():
            self.database_numeros[str(row["Nº"]).strip()] = [row["Agência"], row['Local']]

    def comiple_info(self, limite_horizontal, limite_vertical, file_path='numero.xlsx'):
            def find_and_replace(key, qtd, valor, internet):
                workbook = openpyxl.load_workbook(file_path)
                planilha = workbook.active

                tamanho_local = len(self.locais)+4

                for numero in range(3, tamanho_local):
                    if str(planilha[f"A{numero}"].value).strip() == str(key).strip():
    
                        planilha[f'D{numero}'] = valor
                        planilha[f'B{numero}'] = qtd
                        planilha[f'C{numero}'] = round(float(internet/1024), 3)
                
                workbook.save(file_path)
                workbook.close()

            df = pd.read_excel(file_path, sheet_name="Numeros", skiprows=len(self.locais)+4, usecols=f"A:{limite_horizontal}")
            for agencia, agencia_df in df.groupby("Agência"):
                transform_kb_in_mb = lambda x: x/1024

                find_and_replace(agencia, agencia_df['Agência'].count(), agencia_df['Valor'].sum(), (agencia_df['KB'].sum() + transform_kb_in_mb(agencia_df['KB'].sum())))

    def open_result(self, file_path='numero.xlsx'):
        """
        Abre a planilha formatada e completa.
        """
        os.startfile(file_path)
        

if __name__ == "__main__":
    root = tk.Tk()
    app = Application(root)
    root.mainloop()
